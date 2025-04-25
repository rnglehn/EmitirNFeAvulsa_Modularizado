# report.py

import os
import re
import pandas as pd
from datetime import datetime
from utils import normalize_text, limpar_nome
from openpyxl import Workbook
from openpyxl.styles import Font

def generate_report(dados: dict, classe_escolhida: str, df_pauta: pd.DataFrame) -> str:
    """
    Gera um Excel único com aba "Dados GTA" e um JSON na pasta JSON/,
    trazendo o preço correto da pauta fiscal via merge sobre descrições normalizadas.
    """
    # Data para o relatório
    hoje = datetime.now().strftime('%d/%m/%Y')

    # Cabeçalho GTA
    num_gta = dados.get('numero_gta', '')
    uf      = dados.get('uf', '')
    serie   = dados.get('serie', '')
    finalidade = dados.get('finalidade', '')
    validade   = dados.get('validade', '')

    # 1) Prepara DataFrame das categorias extraídas
    df_cat = pd.DataFrame(dados.get('categorias', []))
    df_cat['Classe'] = classe_escolhida

    # 2) Função de normalização de descrição
    def normalize_desc(s: str) -> str:
        t = str(s).upper()
        # singulariza Bovinos → Bovino
        t = re.sub(r'\bBOVINOS\b', 'BOVINO', t)
        # unifica "+ DE" para casar com pauta que usa "+ DE"
        t = t.replace('+ DE', '+ DE')
        # remove acentos e caracteres supérfluos
        return normalize_text(t)

    # 3) Cria chaves de junção em df_cat
    df_cat['Classe_Normalizada']   = df_cat['Classe'].apply(normalize_text)
    df_cat['Descricao_Normalizada'] = (
        df_cat['especie'].astype(str) + " " +
        df_cat['sexo']  .astype(str) + " " +
        df_cat['faixa'] .astype(str)
    ).apply(normalize_desc)

    # 4) Prepara df_pauta: normaliza descrição e classe
    #    Assumimos que a coluna de descrição se chama 'Descrição'
    if 'Classe' not in df_pauta.columns or 'Descrição' not in df_pauta.columns:
        raise KeyError("df_pauta deve conter colunas 'Classe' e 'Descrição'")
    df_pauta['Classe_Normalizada']   = df_pauta['Classe'].apply(normalize_text)
    df_pauta['Descricao_Normalizada'] = df_pauta['Descrição'].apply(normalize_desc)

    # 5) Localiza e renomeia coluna de preço para 'Valor'
    preco_col = next((c for c in df_pauta.columns if 'valor' in c.lower()), None)
    if not preco_col:
        raise KeyError("Coluna de preço não encontrada na pauta (buscando 'valor').")
    df_pauta = df_pauta.rename(columns={preco_col: 'Valor'})

    # 6) Faz merge para trazer o preço
    df_merge = pd.merge(
        df_cat,
        df_pauta[['Descricao_Normalizada','Classe_Normalizada','Valor']],
        on=['Descricao_Normalizada','Classe_Normalizada'],
        how='left'
    )

    # 7) Monta DataFrame final de produtos
    df_merge['Preço Pauta Fiscal'] = df_merge['Valor']
    df_merge['Data']               = hoje

    df_prod = df_merge[[
        'especie','sexo','faixa','quantidade','Classe','Preço Pauta Fiscal','Data'
    ]].rename(columns={
        'especie':    'Espécie',
        'sexo':       'Sexo',
        'faixa':      'Faixa',
        'quantidade': 'Quantidade'
    })

    # 8) Cria planilha com openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados GTA"
    bold = Font(bold=True)

    # Cabeçalho
    ws['A1'] = "GTA";       ws['A1'].font = bold
    ws['A2'] = "Número";    ws['B2'] = num_gta
    ws['A3'] = "UF";        ws['B3'] = uf
    ws['A4'] = "Série";     ws['B4'] = serie

    # Procedência / Destino
    ws['A6'] = "PROCEDÊNCIA"; ws['A6'].font = bold
    ws['D6'] = "DESTINO";     ws['D6'].font = bold

    # Rótulos fixos em A7–A10
    ws['A7'] = "CPF / CNPJ Procedencia";       ws['A7'].font = bold
    ws['A8'] = "Nome Procedencia";             ws['A8'].font = bold
    ws['A9'] = "Estabelecimento Procedencia";  ws['A9'].font = bold
    ws['A10']= "Municipio Procedencia";        ws['A10'].font = bold

    # Valores em B7–B10
    ws['B7'] = dados.get('cpf_procedencia','')
    ws['B8'] = dados.get('nome_procedencia','')
    ws['B9'] = dados.get('estabelecimento_procedencia','')
    ws['B10']= dados.get('municipio_procedencia','')

    # Rótulos fixos em D7–D10
    ws['D7'] = "CPF / CNPJ Destino";           ws['D7'].font = bold
    ws['D8'] = "Nome Destino";                 ws['D8'].font = bold
    ws['D9'] = "Estabelecimento Destino";      ws['D9'].font = bold
    ws['D10']= "Municipio Destino";            ws['D10'].font = bold

    # Valores em E7–E10
    ws['E7'] = dados.get('cpf_destino','')
    ws['E8'] = dados.get('nome_destino','')
    ws['E9'] = dados.get('estabelecimento_destino','')
    ws['E10']= dados.get('municipio_destino','')

    # Cabeçalho da tabela de produtos (linha 12)
    header_row = 12
    for col_idx, header in enumerate(df_prod.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = bold

    # Preenche produtos
    for row_idx, row in enumerate(df_prod.itertuples(index=False), start=header_row+1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Dados adicionais
    footer = header_row + 1 + len(df_prod) + 1
    ws[f"A{footer}"]   = "DADOS ADICIONAIS";    ws[f"A{footer}"].font = bold
    ws[f"A{footer+1}"] = "Finalidade";           ws[f"B{footer+1}"] = finalidade
    ws[f"A{footer+2}"] = "Validade";             ws[f"B{footer+2}"] = validade

    # Salva Excel
    ts = datetime.now().strftime('%d.%m.%Y_%H-%M-%S')
    base = f"RELATORIO_NFE_GTA_{num_gta}_{limpar_nome(dados.get('nome_procedencia',''))}_{ts}"
    os.makedirs("Relatórios", exist_ok=True)
    excel_path = os.path.join("Relatórios", base + ".xlsx")
    wb.save(excel_path)
    print(f"\n✅ Excel salvo em: {excel_path}")

    # Salva JSON de produtos
    os.makedirs("JSON", exist_ok=True)
    json_path = os.path.join("JSON", base + "_produtos.json")
    df_prod.to_json(json_path, orient='records', force_ascii=False, indent=2)
    print(f"✅ JSON salvo em:  {json_path}")

    return excel_path
